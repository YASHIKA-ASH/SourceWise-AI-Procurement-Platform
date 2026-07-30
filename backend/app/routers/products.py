import csv
import io
from datetime import date, datetime

from fastapi import APIRouter, Depends, File, HTTPException, Response, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..database import get_db
from ..models import BOMComponent, Product
from ..schemas import (
    ComponentCreate,
    ComponentRead,
    ComponentUpdate,
    ProductCreate,
    ProductRead,
    ProductUpdate,
)

router = APIRouter(prefix="/products", tags=["Products and BOM"])


@router.get("", response_model=list[ProductRead])
def list_products(db: Session = Depends(get_db)):
    return db.scalars(select(Product).order_by(Product.name)).all()


@router.post("", response_model=ProductRead, status_code=201)
def create_product(payload: ProductCreate, db: Session = Depends(get_db)):
    if db.scalar(select(Product).where(Product.sku == payload.sku)):
        raise HTTPException(status_code=409, detail="A product with this SKU already exists")
    product = Product(**payload.model_dump())
    db.add(product)
    db.commit()
    db.refresh(product)
    return product


@router.patch("/{product_id}", response_model=ProductRead)
def update_product(product_id: int, payload: ProductUpdate, db: Session = Depends(get_db)):
    product = _get_product(product_id, db)
    values = payload.model_dump(exclude_unset=True)
    if "sku" in values:
        duplicate = db.scalar(
            select(Product).where(Product.sku == values["sku"], Product.id != product_id)
        )
        if duplicate:
            raise HTTPException(status_code=409, detail="A product with this SKU already exists")
    for key, value in values.items():
        setattr(product, key, value)
    db.commit()
    db.refresh(product)
    return product


@router.delete("/{product_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_product(product_id: int, db: Session = Depends(get_db)):
    product = _get_product(product_id, db)
    db.delete(product)
    db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.get("/{product_id}/bom", response_model=list[ComponentRead])
def get_bom(product_id: int, db: Session = Depends(get_db)):
    _get_product(product_id, db)
    return db.scalars(
        select(BOMComponent)
        .where(BOMComponent.product_id == product_id)
        .order_by(BOMComponent.required_delivery_date, BOMComponent.part_name)
    ).all()


@router.post("/{product_id}/components", response_model=ComponentRead, status_code=201)
def add_component(product_id: int, payload: ComponentCreate, db: Session = Depends(get_db)):
    _get_product(product_id, db)
    component = BOMComponent(product_id=product_id, **payload.model_dump())
    db.add(component)
    db.commit()
    db.refresh(component)
    return component


@router.patch("/{product_id}/components/{component_id}", response_model=ComponentRead)
def update_component(
    product_id: int,
    component_id: int,
    payload: ComponentUpdate,
    db: Session = Depends(get_db),
):
    _get_product(product_id, db)
    component = db.get(BOMComponent, component_id)
    if not component or component.product_id != product_id:
        raise HTTPException(status_code=404, detail="Component not found")
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(component, key, value)
    db.commit()
    db.refresh(component)
    return component


@router.delete("/{product_id}/components/{component_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_component(product_id: int, component_id: int, db: Session = Depends(get_db)):
    _get_product(product_id, db)
    component = db.get(BOMComponent, component_id)
    if not component or component.product_id != product_id:
        raise HTTPException(status_code=404, detail="Component not found")
    db.delete(component)
    db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.post("/{product_id}/bom/upload")
async def upload_bom(product_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    _get_product(product_id, db)
    filename = (file.filename or "").lower()
    raw = await file.read()
    rows: list[dict] = []

    if filename.endswith(".csv"):
        text = raw.decode("utf-8-sig")
        rows = list(csv.DictReader(io.StringIO(text)))
    elif filename.endswith(".xlsx"):
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(sheet.iter_rows())]
        for values in sheet.iter_rows(values_only=True):
            rows.append(dict(zip(headers, values)))
    else:
        raise HTTPException(status_code=400, detail="Upload a CSV or XLSX file")

    created = []
    errors = []
    for index, row in enumerate(rows, start=2):
        try:
            normalized = {str(key).strip().lower().replace(" ", "_"): value for key, value in row.items()}
            component = BOMComponent(
                product_id=product_id,
                part_name=str(normalized["part_name"]).strip(),
                category=str(normalized.get("category") or "Uncategorized").strip(),
                required_quantity=float(normalized["required_quantity"]),
                current_inventory=float(normalized.get("current_inventory") or 0),
                reserved_inventory=float(normalized.get("reserved_inventory") or 0),
                safety_stock=float(normalized.get("safety_stock") or 0),
                default_minimum_order_quantity=float(normalized.get("minimum_order_quantity") or 1),
                required_delivery_date=_parse_date(normalized["required_delivery_date"]),
                is_critical=_parse_bool(normalized.get("is_critical", True)),
            )
            db.add(component)
            created.append(component)
        except Exception as exc:
            errors.append({"row": index, "error": str(exc)})

    if errors:
        db.rollback()
        raise HTTPException(status_code=422, detail={"message": "BOM validation failed", "errors": errors})

    db.commit()
    return {"created_components": len(created), "filename": file.filename}


def _get_product(product_id: int, db: Session) -> Product:
    product = db.get(Product, product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return product


def _parse_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value).strip())


def _parse_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"true", "1", "yes", "y"}
